"""
Fixed Multiplier Mode 계산 검증 문서 생성기

PRD 요구사항에 따른 계산 공식:
1. Target Position Size = Multiplier × Master Position Size
2. Effective Leverage = min(Master Leverage, Copier Leverage Limit)
3. Target Notional = Target Position Size × Entry Price
4. Required Margin = Target Notional / Effective Leverage
5. Validation: Required Margin <= Available Balance
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

class CalculationVerificationGenerator:
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # 기본 시트 제거
        
        # 스타일 정의
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.input_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        self.formula_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_formula_reference_sheet(self):
        """수식 참조 시트 생성"""
        ws = self.wb.create_sheet("📐 계산 공식")
        
        # 제목
        ws['A1'] = "Fixed Multiplier Mode 계산 공식 참조"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        # PRD 정보
        ws['A3'] = "PRD 버전:"
        ws['B3'] = "Fixed Multiplier Mode v2"
        ws['A4'] = "문서 생성일:"
        ws['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 계산 공식
        formulas = [
            ["번호", "항목", "공식", "설명"],
            [1, "Target Position Size", "Multiplier × Master Position Size", "복사될 포지션 크기 (USDT)"],
            [2, "Effective Leverage", "MIN(Master Leverage, Copier Leverage Limit)", "실제 적용될 레버리지"],
            [3, "Entry Price", "Master Entry Price (± Slippage)", "진입 가격 (슬리피지 고려)"],
            [4, "Target Notional", "Target Position Size × Entry Price", "명목 가치"],
            [5, "Required Margin", "Target Notional / Effective Leverage", "필요한 마진"],
            [6, "Validation", "Required Margin ≤ Available Balance", "주문 실행 가능 여부"],
            [7, "Order Result", "IF(Validation = TRUE, 'Success', 'Failed: Insufficient Margin')", "주문 결과"]
        ]
        
        row = 6
        for formula in formulas:
            for col, value in enumerate(formula, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                if row == 6:
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row += 1
        
        # 제약 조건
        ws['A' + str(row + 2)] = "제약 조건 (Constraints):"
        ws['A' + str(row + 2)].font = Font(bold=True, size=12)
        
        constraints = [
            ["항목", "최소값", "최대값", "단위"],
            ["Multiplier", "0.01", "100", "x"],
            ["Master Leverage", "1", "100", "x (Category별 제한)"],
            ["Copier Leverage Limit", "1", "100", "x (Category별 제한)"],
            ["Category 1 Max Leverage", "-", "100", "x"],
            ["Category 2 Max Leverage", "-", "50", "x"],
            ["Category 3~8 Max Leverage", "-", "20", "x"]
        ]
        
        row = row + 4
        for constraint in constraints:
            for col, value in enumerate(constraint, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                if row == row:
                    cell.fill = self.header_fill
                    cell.font = self.header_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 40
    
    def create_test_cases_sheet(self):
        """테스트 케이스 시트 생성 (수식 포함)"""
        ws = self.wb.create_sheet("✅ 테스트 케이스")
        
        # 제목
        ws['A1'] = "Fixed Multiplier Mode 계산 검증 테스트 케이스"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:O1')
        
        # 헤더
        headers = [
            "TC ID", "시나리오", 
            "Master Position\n(USDT)", "Multiplier\n(x)", "Master\nLeverage", "Copier Leverage\nLimit", "Entry Price\n(USDT)", "Available\nBalance (USDT)",
            "Target Size\n(USDT)", "Effective\nLeverage", "Target\nNotional", "Required\nMargin",
            "Validation", "Order Result", "비고"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 테스트 케이스 데이터
        test_cases = [
            # TC ID, 시나리오, Master Pos, Multiplier, Master Lev, Copier Lev Limit, Entry Price, Available Balance
            ["TC-001", "일반 케이스 (2.5x)", 100, 2.5, 10, 10, 50000, 2000],
            ["TC-002", "최소 Multiplier (0.01x)", 1000, 0.01, 10, 10, 50000, 1000],
            ["TC-003", "최대 Multiplier (100x)", 100, 100, 10, 10, 50000, 100000],
            ["TC-004", "Leverage Limit 우선 적용", 100, 5, 50, 20, 50000, 20000],
            ["TC-005", "마진 부족 (실패)", 100, 10, 10, 10, 50000, 100],
            ["TC-006", "마진 딱 맞음 (성공)", 100, 5, 10, 10, 50000, 2500],
            ["TC-007", "소수점 Multiplier (1.5x)", 200, 1.5, 10, 10, 50000, 5000],
            ["TC-008", "높은 Entry Price", 100, 2, 10, 10, 100000, 5000],
            ["TC-009", "낮은 Leverage (1x)", 1000, 5, 1, 1, 50000, 300000],
            ["TC-010", "Category 1 (100x Limit)", 50, 10, 100, 100, 50000, 30000],
            ["TC-011", "Category 2 (50x Limit)", 50, 10, 50, 50, 50000, 15000],
            ["TC-012", "Category 3-8 (20x Limit)", 50, 10, 20, 20, 50000, 15000],
            ["TC-013", "극단값: 0.01x, Price 1", 10000, 0.01, 10, 10, 1, 100],
            ["TC-014", "극단값: 100x, Max Leverage", 10, 100, 100, 100, 50000, 100000],
            ["TC-015", "경계값: Multiplier 0.01x", 500, 0.01, 10, 10, 50000, 100],
            ["TC-016", "경계값: Multiplier 100x", 10, 100, 10, 10, 50000, 10000],
            ["TC-017", "소수점 Position Size", 123.45, 2.5, 10, 10, 50000, 5000],
            ["TC-018", "Master Leverage > Copier Limit", 100, 3, 75, 25, 50000, 10000],
            ["TC-019", "Master Leverage < Copier Limit", 100, 3, 15, 50, 50000, 10000],
            ["TC-020", "마진 거의 부족 (99%)", 100, 10, 10, 10, 50000, 499],
        ]
        
        row = 4
        for tc in test_cases:
            # 입력값
            for col, value in enumerate(tc, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col <= 8:  # 입력 필드
                    cell.fill = self.input_fill
            
            # 수식 셀 (I~L열: Target Size, Effective Leverage, Target Notional, Required Margin)
            # I: Target Size = Master Position × Multiplier
            ws.cell(row=row, column=9, value=f"=C{row}*D{row}")
            ws.cell(row=row, column=9).fill = self.formula_fill
            ws.cell(row=row, column=9).border = self.border
            ws.cell(row=row, column=9).number_format = '0.00'
            
            # J: Effective Leverage = MIN(Master Leverage, Copier Leverage Limit)
            ws.cell(row=row, column=10, value=f"=MIN(E{row},F{row})")
            ws.cell(row=row, column=10).fill = self.formula_fill
            ws.cell(row=row, column=10).border = self.border
            ws.cell(row=row, column=10).number_format = '0.00'
            
            # K: Target Notional = Target Size × Entry Price
            ws.cell(row=row, column=11, value=f"=I{row}*G{row}")
            ws.cell(row=row, column=11).fill = self.formula_fill
            ws.cell(row=row, column=11).border = self.border
            ws.cell(row=row, column=11).number_format = '0.00'
            
            # L: Required Margin = Target Notional / Effective Leverage
            ws.cell(row=row, column=12, value=f"=K{row}/J{row}")
            ws.cell(row=row, column=12).fill = self.formula_fill
            ws.cell(row=row, column=12).border = self.border
            ws.cell(row=row, column=12).number_format = '0.00'
            
            # M: Validation = Required Margin <= Available Balance
            ws.cell(row=row, column=13, value=f"=IF(L{row}<=H{row},\"PASS\",\"FAIL\")")
            ws.cell(row=row, column=13).border = self.border
            ws.cell(row=row, column=13).alignment = Alignment(horizontal='center', vertical='center')
            
            # N: Order Result
            ws.cell(row=row, column=14, value=f"=IF(M{row}=\"PASS\",\"Success\",\"Failed: Insufficient Margin\")")
            ws.cell(row=row, column=14).border = self.border
            ws.cell(row=row, column=14).alignment = Alignment(horizontal='center', vertical='center')
            
            # O: 비고 (계산 상세)
            remark = f"=CONCATENATE(\"Target: \",TEXT(I{row},\"0.00\"),\" USDT | Margin: \",TEXT(L{row},\"0.00\"),\"/\",TEXT(H{row},\"0.00\"),\" USDT\")"
            ws.cell(row=row, column=15, value=remark)
            ws.cell(row=row, column=15).border = self.border
            ws.cell(row=row, column=15).alignment = Alignment(horizontal='left', vertical='center')
            
            row += 1
        
        # 조건부 서식 (Pass/Fail)
        for r in range(4, row):
            validation_cell = ws.cell(row=r, column=13)
            result_cell = ws.cell(row=r, column=14)
            # Excel 조건부 서식은 openpyxl로 직접 적용하기 어려우므로, 
            # 나중에 수동으로 적용하거나 별도 라이브러리 사용
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 13
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 25
        ws.column_dimensions['O'].width = 40
        
        # 행 높이
        ws.row_dimensions[3].height = 40
    
    def create_edge_cases_sheet(self):
        """엣지 케이스 시트 생성"""
        ws = self.wb.create_sheet("⚠️ 엣지 케이스")
        
        # 제목
        ws['A1'] = "Fixed Multiplier Mode 엣지 케이스 검증"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:O1')
        
        # 헤더 (동일)
        headers = [
            "TC ID", "엣지 케이스 시나리오", 
            "Master Position\n(USDT)", "Multiplier\n(x)", "Master\nLeverage", "Copier Leverage\nLimit", "Entry Price\n(USDT)", "Available\nBalance (USDT)",
            "Target Size\n(USDT)", "Effective\nLeverage", "Target\nNotional", "Required\nMargin",
            "Validation", "Order Result", "예상 동작"
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 엣지 케이스 데이터
        edge_cases = [
            ["EDGE-001", "Multiplier 범위 초과 (101x)", 100, 101, 10, 10, 50000, 10000, "입력 차단 또는 에러"],
            ["EDGE-002", "Multiplier 범위 미달 (0.009x)", 100, 0.009, 10, 10, 50000, 1000, "입력 차단 또는 에러"],
            ["EDGE-003", "Position Size = 0", 0, 5, 10, 10, 50000, 1000, "주문 실패 (Size 0)"],
            ["EDGE-004", "Entry Price = 0", 100, 5, 10, 10, 0, 1000, "주문 실패 (Price 0)"],
            ["EDGE-005", "Available Balance = 0", 100, 5, 10, 10, 50000, 0, "주문 실패 (잔고 부족)"],
            ["EDGE-006", "Leverage = 0", 100, 5, 0, 0, 50000, 10000, "시스템 에러 (Leverage 0)"],
            ["EDGE-007", "매우 작은 Target Size (0.01 USDT)", 0.01, 1, 10, 10, 50000, 100, "최소 주문 크기 확인"],
            ["EDGE-008", "매우 큰 Target Size (1M USDT)", 10000, 100, 10, 10, 50000, 10000000, "최대 주문 크기 확인"],
            ["EDGE-009", "소수점 정밀도 (0.123456x)", 100, 0.123456, 10, 10, 50000, 1000, "소수점 처리 확인"],
            ["EDGE-010", "음수 Multiplier (-1x)", 100, -1, 10, 10, 50000, 1000, "입력 차단"],
        ]
        
        row = 4
        for tc in edge_cases:
            # TC ID, 시나리오
            ws.cell(row=row, column=1, value=tc[0]).border = self.border
            ws.cell(row=row, column=2, value=tc[1]).border = self.border
            
            # 입력값 (범위 검증을 위해 일부러 잘못된 값 포함)
            for col in range(3, 9):
                cell = ws.cell(row=row, column=col, value=tc[col-1])
                cell.fill = self.input_fill
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 수식 (에러가 발생할 수 있으므로 IFERROR 사용)
            # I: Target Size
            ws.cell(row=row, column=9, value=f"=IFERROR(C{row}*D{row},\"ERROR\")")
            ws.cell(row=row, column=9).fill = self.formula_fill
            ws.cell(row=row, column=9).border = self.border
            
            # J: Effective Leverage
            ws.cell(row=row, column=10, value=f"=IFERROR(MIN(E{row},F{row}),\"ERROR\")")
            ws.cell(row=row, column=10).fill = self.formula_fill
            ws.cell(row=row, column=10).border = self.border
            
            # K: Target Notional
            ws.cell(row=row, column=11, value=f"=IFERROR(I{row}*G{row},\"ERROR\")")
            ws.cell(row=row, column=11).fill = self.formula_fill
            ws.cell(row=row, column=11).border = self.border
            
            # L: Required Margin
            ws.cell(row=row, column=12, value=f"=IFERROR(K{row}/J{row},\"ERROR\")")
            ws.cell(row=row, column=12).fill = self.formula_fill
            ws.cell(row=row, column=12).border = self.border
            
            # M: Validation
            ws.cell(row=row, column=13, value=f"=IF(OR(ISERROR(L{row}),L{row}=\"ERROR\"),\"ERROR\",IF(L{row}<=H{row},\"PASS\",\"FAIL\"))")
            ws.cell(row=row, column=13).border = self.border
            
            # N: Order Result
            ws.cell(row=row, column=14, value=f"=IF(M{row}=\"PASS\",\"Success\",IF(M{row}=\"ERROR\",\"System Error\",\"Failed\"))")
            ws.cell(row=row, column=14).border = self.border
            
            # O: 예상 동작
            ws.cell(row=row, column=15, value=tc[8]).border = self.border
            ws.cell(row=row, column=15).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            row += 1
        
        # 열 너비
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 13
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 20
        ws.column_dimensions['O'].width = 30
        
        ws.row_dimensions[3].height = 40
    
    def create_summary_sheet(self):
        """요약 시트 생성"""
        ws = self.wb.create_sheet("📊 검증 요약", 0)  # 첫 번째 시트로
        
        # 제목
        ws['A1'] = "Fixed Multiplier Mode 계산 검증 요약"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "문서 정보"
        ws['A3'].font = Font(bold=True, size=12)
        
        info = [
            ["생성일시:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["PRD 버전:", "Fixed Multiplier Mode v2"],
            ["검증 항목:", "Position Size, Leverage, Margin 계산"],
            ["총 테스트 케이스:", "20개 (일반) + 10개 (엣지)"]
        ]
        
        row = 4
        for item in info:
            ws.cell(row=row, column=1, value=item[0]).font = Font(bold=True)
            ws.cell(row=row, column=2, value=item[1])
            row += 1
        
        # 사용 방법
        ws['A' + str(row + 2)] = "📖 사용 방법"
        ws['A' + str(row + 2)].font = Font(bold=True, size=12)
        
        instructions = [
            "1. '✅ 테스트 케이스' 시트에서 입력값(회색 셀)을 수정하여 다양한 시나리오 테스트",
            "2. 계산 결과(노란색 셀)는 자동으로 업데이트됨",
            "3. 'Validation' 열에서 PASS/FAIL 확인",
            "4. 'Order Result' 열에서 예상 주문 결과 확인",
            "5. '⚠️ 엣지 케이스' 시트에서 경계값 및 오류 케이스 확인",
            "6. '📐 계산 공식' 시트에서 PRD 공식 참조"
        ]
        
        row = row + 3
        for instruction in instructions:
            ws.cell(row=row, column=1, value=instruction)
            ws.merge_cells(f'A{row}:D{row}')
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row += 1
        
        # 주요 공식
        ws['A' + str(row + 2)] = "🔢 주요 계산 공식"
        ws['A' + str(row + 2)].font = Font(bold=True, size=12)
        
        formulas_summary = [
            ["1. Target Position Size", "= Multiplier × Master Position Size"],
            ["2. Effective Leverage", "= MIN(Master Leverage, Copier Leverage Limit)"],
            ["3. Target Notional", "= Target Position Size × Entry Price"],
            ["4. Required Margin", "= Target Notional / Effective Leverage"],
            ["5. Order Success", "= Required Margin ≤ Available Balance"]
        ]
        
        row = row + 3
        for formula in formulas_summary:
            ws.cell(row=row, column=1, value=formula[0]).font = Font(bold=True)
            ws.cell(row=row, column=2, value=formula[1])
            ws.merge_cells(f'B{row}:D{row}')
            row += 1
        
        # 제약 조건
        ws['A' + str(row + 2)] = "⚠️ 제약 조건"
        ws['A' + str(row + 2)].font = Font(bold=True, size=12)
        
        constraints_summary = [
            ["Multiplier 범위:", "0.01x ~ 100x"],
            ["소수점 정밀도:", "2자리 권장 (더 높은 정밀도 허용)"],
            ["마진 부족 시:", "자동 스케일 다운 없이 주문 실패"],
            ["Leverage 제한:", "Category별 상이 (1: 100x, 2: 50x, 3~8: 20x)"]
        ]
        
        row = row + 3
        for constraint in constraints_summary:
            ws.cell(row=row, column=1, value=constraint[0]).font = Font(bold=True)
            ws.cell(row=row, column=2, value=constraint[1])
            ws.merge_cells(f'B{row}:D{row}')
            row += 1
        
        # 열 너비
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
    
    def generate(self, output_path):
        """검증 문서 생성"""
        print("📊 Fixed Multiplier Mode 계산 검증 문서 생성 중...")
        
        self.create_summary_sheet()
        self.create_formula_reference_sheet()
        self.create_test_cases_sheet()
        self.create_edge_cases_sheet()
        
        self.wb.save(output_path)
        print(f"✅ 검증 문서 생성 완료: {output_path}")
        
        return output_path


def main():
    generator = CalculationVerificationGenerator()
    
    output_dir = "output/fixed_multiplier_integrated"
    os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"{output_dir}/FixedMultiplier_Calculation_Verification_{timestamp}.xlsx"
    
    generator.generate(output_file)
    
    print("\n" + "="*70)
    print("🎉 계산 검증 문서 생성 완료!")
    print("="*70)
    print(f"\n📁 파일 위치: {output_file}")
    print("\n📋 포함된 시트:")
    print("   1. 📊 검증 요약 - 사용 방법 및 주요 정보")
    print("   2. 📐 계산 공식 - PRD 기반 공식 참조")
    print("   3. ✅ 테스트 케이스 - 20개 일반 시나리오 (수식 자동 계산)")
    print("   4. ⚠️ 엣지 케이스 - 10개 경계값/오류 시나리오")
    print("\n💡 사용법:")
    print("   - 회색 셀: 입력값 (수정 가능)")
    print("   - 노란색 셀: 계산 결과 (자동 업데이트)")
    print("   - Validation 열에서 PASS/FAIL 확인")
    print("="*70)


if __name__ == "__main__":
    main()
