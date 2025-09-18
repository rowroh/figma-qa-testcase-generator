#!/usr/bin/env python3
"""
🎯 Figma 테스트케이스 생성 MCP 서버

🏆 업데이트: Trading Competition = 랭킹 시스템 (NOT 티어 시스템)
🌟 추가: User VIP System = 티어 시스템 (멤버십 등급 관리)

이 MCP 서버는 다음 도구들을 제공합니다:
- parse_figma_url: Figma URL 파싱
- fetch_figma_data: Figma API 데이터 가져오기
- extract_requirements: 요구사항 추출 (랭킹 + VIP 티어 시스템 키워드 포함)
- generate_testcase: 테스트케이스 생성 (랭킹 + VIP 티어 기반 분류)
- save_to_excel: Excel 저장
- process_figma_link: 전체 프로세스 실행
- enhanced_figma_analysis: 향상된 분석 (랭킹 + VIP 티어 시스템 패턴 포함)
"""

import json
import asyncio
import sys
from typing import Any, Sequence
import requests
import pandas as pd
from datetime import datetime
from urllib.parse import urlparse, parse_qs
import os
from dotenv import load_dotenv

# MCP 관련 import
try:
    from mcp.server.models import InitializationOptions
    from mcp.server import NotificationOptions, Server
    from mcp.types import (
        CallToolRequest,
        CallToolResult,
        ListToolsRequest,
        ListToolsResult,
        Tool,
        TextContent,
        ImageContent,
        EmbeddedResource,
    )
    MCP_AVAILABLE = True
except ImportError:
    print("❌ MCP 라이브러리가 설치되지 않았습니다.")
    print("설치: pip install mcp")
    MCP_AVAILABLE = False

# 환경변수 로드
load_dotenv()

class FigmaMCPServer:
    def __init__(self):
        self.figma_token = os.getenv("FIGMA_TOKEN")
        if not self.figma_token:
            raise ValueError("FIGMA_TOKEN 환경변수가 설정되지 않았습니다.")
    
    def parse_figma_url(self, figma_url: str) -> dict:
        """Figma URL에서 파일 ID와 노드 ID 추출"""
        try:
            parsed_url = urlparse(figma_url)
            path_parts = parsed_url.path.split('/')
            
            # 파일 ID 추출 (file 또는 design 경로 지원)
            file_id = None
            if 'file' in path_parts:
                file_index = path_parts.index('file')
                if file_index + 1 < len(path_parts):
                    file_id = path_parts[file_index + 1]
            elif 'design' in path_parts:
                design_index = path_parts.index('design')
                if design_index + 1 < len(path_parts):
                    file_id = path_parts[design_index + 1]
            
            # 노드 ID 추출
            query_params = parse_qs(parsed_url.query)
            node_id = None
            if 'node-id' in query_params:
                node_id = query_params['node-id'][0].replace('%3A', ':')
            
            return {
                "success": True,
                "file_id": file_id,
                "node_id": node_id,
                "url": figma_url
            }
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def fetch_figma_data(self, file_id: str, node_id: str = None) -> dict:
        """Figma API에서 데이터 가져오기"""
        headers = {"X-Figma-Token": self.figma_token}
        
        try:
            if node_id:
                # 노드 ID 형식 변환 (2-4 -> 2:4)
                node_id_formatted = node_id.replace('-', ':')
                url = f"https://api.figma.com/v1/files/{file_id}/nodes?ids={node_id_formatted}"
                response = requests.get(url, headers=headers)
                response.raise_for_status()
                data = response.json()
                
                if 'nodes' in data and node_id_formatted in data['nodes']:
                    return {
                        "success": True,
                        "data": {
                            'document': {
                                'children': [data['nodes'][node_id_formatted]['document']]
                            }
                        }
                    }
                else:
                    # 특정 노드를 찾지 못한 경우 전체 파일 가져오기
                    url = f"https://api.figma.com/v1/files/{file_id}"
                    response = requests.get(url, headers=headers)
                    response.raise_for_status()
                    return {
                        "success": True,
                        "data": response.json(),
                        "note": f"Node {node_id_formatted} not found, returning full file"
                    }
            else:
                url = f"https://api.figma.com/v1/files/{file_id}"
                response = requests.get(url, headers=headers)
                response.raise_for_status()
                return {
                    "success": True,
                    "data": response.json()
                }
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def extract_requirements(self, figma_data: dict) -> dict:
        """Figma 데이터에서 요구사항 추출"""
        requirements = []
        
        def traverse_nodes(nodes):
            if isinstance(nodes, list):
                for node in nodes:
                    traverse_nodes(node)
            elif isinstance(nodes, dict):
                # 텍스트 노드에서 추출
                if nodes.get('type') == 'TEXT' and 'characters' in nodes:
                    text = nodes['characters'].strip()
                    if self._is_requirement_text(text):
                        requirements.append({
                            'source': 'text_node',
                            'text': text,
                            'node_name': nodes.get('name', ''),
                            'node_id': nodes.get('id', '')
                        })
                
                # 프레임/컴포넌트 이름에서 추출
                elif nodes.get('type') in ['FRAME', 'COMPONENT', 'INSTANCE']:
                    name = nodes.get('name', '')
                    if self._is_requirement_text(name):
                        requirements.append({
                            'source': 'frame_name',
                            'text': name,
                            'node_name': name,
                            'node_id': nodes.get('id', ''),
                            'type': nodes.get('type', '')
                        })
                
                # 자식 노드 재귀 탐색
                if 'children' in nodes:
                    traverse_nodes(nodes['children'])
        
        if figma_data and 'document' in figma_data:
            traverse_nodes(figma_data['document'].get('children', []))
        
        # 중복 제거
        unique_requirements = self._deduplicate_requirements(requirements)
        
        return {
            "success": True,
            "requirements": unique_requirements,
            "count": len(unique_requirements)
        }
    
    def _is_requirement_text(self, text: str) -> bool:
        """텍스트가 요구사항인지 판단"""
        if not text or len(text) < 3 or len(text) > 1000:
            return False
        
        requirement_keywords = [
            '기능', '요구사항', '사용자', '시스템', '화면', '페이지', '버튼',
            '클릭', '선택', '입력', '검색', '필터', '정렬', '스크롤',
            '로그인', '회원가입', '로그아웃', '프로필', '설정', '알림',
            '목록', '리스트', '카드', '메뉴', '탭', '모달', '팝업',
            '등록', '수정', '삭제', '추가', '업데이트', '동기화',
            '거래', '주문', '보유', '자산', '포트폴리오', '접근성', '표시', '대기',
            '자산관리', '잔고', '잔액', '총자산', '수익률', '계좌', '지갑',
            '입금', '출금', '이체', '매수', '매도', '체결', '미체결', '취소',
            '차트', '그래프', '통계', '분석', '리포트', '히스토리', '거래내역',
            '대시보드', '새로고침', '실시간', '진행중', '완료', '실패', '승인',
            # 피그마에서 추출된 새로운 키워드들 (한국어)
            '출금한도', '한도', '인증', '검증', '보호', '손실', '평균', '단가',
            '수익', '자동매수', '주소록', '주소', '확인', '네트워크', '블록체인',
            '시스템점검', '점검', '내역', '포지션', '손익', '실현', '보류',
            '상태', '유지보수', '전환', '거래정지', '무기한', '스왑', '트리거',
            '실행', '대기중', '부족', '한계', '제한', '활성화', '업그레이드',
            'login', 'signup', 'profile', 'setting', 'notification',
            'search', 'filter', 'sort', 'upload', 'download',
            'button', 'click', 'tap', 'swipe', 'scroll',
            'spot', 'holdings', 'accessibility', 'display', 'pending', 'order',
            'asset', 'portfolio', 'wallet', 'balance', 'total', 'deposit',
            'withdrawal', 'transfer', 'transaction', 'buy', 'sell', 'trade',
            'exchange', 'swap', 'profit', 'loss', 'chart', 'graph', 'analytics',
            'report', 'history', 'dashboard', 'refresh', 'realtime', 'processing',
            # 피그마에서 추출된 새로운 키워드들 (영어)
            'funds', 'available', 'APR', 'withdraw', 'limit', 'verification',
            'protection', 'average', 'cost', 'recurring', 'address', 'book',
            'confirm', 'network', 'blockchain', 'maintenance', 'position',
            'realized', 'cancel', 'status', 'convertible', 'tradable', 'insufficient',
            'icon', 'unfilled', 'filled', 'media', 'radio', 'document',
            'container', 'collapse', 'detail', 'trend', 'mini', 'graphic',
            'boosted', 'protected', 'effective', 'contracts', 'trigger',
            'execution', 'awaiting', 'perpetual', 'activate', 'upgrades',
            'instantly', 'additionally', 'tradeable', 'non-tradable',
            # 노드 추출 기반 보강 키워드
            'market order', 'trigger order', 'order preview', 'order confirmation', 'order form',
            'trade settings', 'positions', 'available funds', 'schedule order', 'trigger time',
            'cancel order', 'unrealized p&l', 'take profit', 'stop loss', 'close position',
            'estimated total value', 'funding fee', 'funding payment', 'auto-deleveraging',
            'trading limit tier', 'perpetual swap', 'picture-in-picture', 'order value', 'max order value',
            'leverage', 'multi-position mode', 'open positions', 'view holdings', 'latest trade', 'top traders',
            # Curated from latest Figma screen analysis
            'tab', 'PnL', 'funding', 'trending', 'favorites',
            'volume', 'symbol', 'banner', 'badge', 'calendar',
            'long', 'short', 'perp', 'market', 'trading',
            # Curated from second Figma screen analysis
            'feed', 'news', 'insights', 'crypto', 'price', 'assets',
            'schedule', 'vip', 'vipstatus',
            'social', 'events', 'economic', 'government', 'user', 'menu',
            # Curated from third Figma screen analysis (Earn/Staking features)
            'stake', 'staked', 'pool', 'rewards',
            'earnings', 'launchpool', 'convert', 'auction',
            'sparks', 'reward', 'flipster', 'pixel', 'ton',
            # Curated from fourth Figma screen analysis (Promotion/Referral hub)
            'claimed', 'hub', 'bonus', 'promotion', 'promotions', 'tasks',
            'complete', 'completed', 'learn', 'identity', 'first', 'referee',
            'referral', 'link',
            # Curated from fifth Figma screen analysis (Notifications/Settings)
            'notifications', 'liquidation', 'liquidated',
            'system', 'notified', 'alerts',
            'action', 'announcement', 'currency', 'warning',
            'reached', 'initial', 'avoid', 'successful', 'amount', 'application',
            '경우', '특정', '모든',
            # Curated from sixth Figma screen analysis (Comprehensive dashboard)
            'program', 'league', 'tier', 'level', 'benefits',
            'day', 'time', 'share', 'empty',
            'secondary', 'choice', 'my', 'basic', 'logomark', 'cropped',
            # Curated from seventh Figma screen analysis (Login/Signup/Registration)
            'log', 'password', 'email', 'input', 'hint',
            'placeholder', 'checkbox', 'terms', 'privacy', 'notice',
            'create', 'code', 'zero', 'fast', 'pairs', 'data',
            'services', 'sso', 'body', 'title',
            # 🏆 UPDATED: Trading Competition은 랭킹 시스템 (NOT 티어 시스템)
            # 랭킹 시스템 키워드 (한국어)
            '랭킹', '순위', '리더보드', '1위', '2위', '3위', '순위표', '등수',
            '상위권', '순위권', '꼴지', '순위변동', '순위상승', '순위하락',
            '경쟁', '경쟁자', '대회', '참가', '참가자', '우승', '우승자',
            '마일스톤', '달성', '목표', '진행률', '보상', '상금',
            '최종순위', '순위별보상', '차등보상', '참가보상',
            # 랭킹 시스템 키워드 (영어)
            'ranking', 'rank', 'leaderboard', '1st', '2nd', '3rd', 'first', 'second', 'third',
            'position', 'standing', 'top', 'bottom', 'rank up', 'rank down',
            'competition', 'competitor', 'participant', 'winner', 'champion',
            'milestone', 'achievement', 'goal', 'progress', 'reward', 'prize',
            'final rank', 'rank-based', 'tier-free', 'dynamic ranking',
            # 🌟 ADDED: VIP 티어 시스템 (User Membership)
            # VIP 티어 시스템 키워드 (한국어)
            'VIP', 'SVIP', 'vip', 'svip', '티어', '등급', '멤버십', '회원등급', '사용자등급',
            '베이직', '실버', '골드', '플래티넘', '승급', '강등', '업그레이드', '다운그레이드',
            '혜택', '특권', '할인', '수수료할인', '전용서비스', '우대서비스', '프리미엄',
            '진행률', '달성률', '요구사항', '조건', '거래량기준', '수수료기준', '보유기간',
            '티어배지', '등급표시', '멤버십카드', '승급진행률', '다음등급', '현재등급',
            # VIP 티어 시스템 키워드 (영어)  
            'vip', 'svip', 'premium', 'elite', 'exclusive', 'tier', 'grade', 'level', 'membership', 'status',
            'basic', 'silver', 'gold', 'platinum', 'upgrade', 'downgrade', 'promotion', 'demotion',
            'benefit', 'privilege', 'discount', 'fee discount', 'exclusive service', 'premium service',
            'progress', 'achievement', 'requirement', 'criteria', 'trading volume', 'fee threshold', 'tenure',
            'tier badge', 'grade display', 'membership card', 'upgrade progress', 'next tier', 'current tier'
        ]
        
        exclude_keywords = [
            'px', 'pt', 'rem', 'color', 'font', 'weight', 'size',
            'margin', 'padding', 'border', 'shadow', 'opacity'
        ]
        
        text_lower = text.lower()
        if any(exclude in text_lower for exclude in exclude_keywords):
            return False
        
        return any(keyword in text for keyword in requirement_keywords)
    
    def _deduplicate_requirements(self, requirements: list) -> list:
        """중복 요구사항 제거"""
        seen = set()
        unique_requirements = []
        
        for req in requirements:
            text_hash = req['text'].strip().lower()
            if text_hash not in seen:
                seen.add(text_hash)
                unique_requirements.append(req)
        
        return unique_requirements
    
    def generate_testcase_structure(self, requirement: dict, test_type: str = None) -> dict:
        """테스트케이스 구조 생성 (AI 없이 기본 템플릿)"""
        req_text = requirement.get('text', '')
        
        # 기본 카테고리 분류 (업데이트: 랭킹 시스템 포함)
        category = "일반기능"
        if any(keyword in req_text.lower() for keyword in ['로그인', 'login', '인증', 'auth']):
            category = "사용자인증"
        elif any(keyword in req_text.lower() for keyword in ['프로필', 'profile', '사용자정보']):
            category = "프로필관리"
        elif any(keyword in req_text.lower() for keyword in ['알림', 'notification', '푸시']):
            category = "알림시스템"
        elif any(keyword in req_text.lower() for keyword in ['검색', 'search', '필터', 'filter']):
            category = "검색기능"
        elif any(keyword in req_text.lower() for keyword in ['랭킹', 'ranking', '순위', 'rank', '리더보드', 'leaderboard', '대회', 'competition']):
            category = "랭킹시스템"
        elif any(keyword in req_text.lower() for keyword in ['vip', 'svip', '티어', 'tier', '등급', 'grade', '멤버십', 'membership', '승급', 'upgrade']):
            category = "VIP티어시스템"
        
        # 우선순위 분류 (업데이트: 랭킹 시스템 고려)
        priority = "P2"
        if any(keyword in req_text.lower() for keyword in ['로그인', '회원가입', '결제', '보안', '랭킹', 'ranking', '순위', 'rank', '대회', 'competition', 'vip', 'svip', '티어', 'tier', '멤버십', 'membership']):
            priority = "P1"
        elif any(keyword in req_text.lower() for keyword in ['설정', '프로필', '알림', '리더보드', 'leaderboard', '마일스톤', 'milestone', '등급', 'grade', '혜택', 'benefit']):
            priority = "P2"
        else:
            priority = "P3"
        
        # 기본 테스트케이스 구조
        testcase = {
            "카테고리": category,
            "제목": f"{req_text} 기능 테스트",
            "참조 ID": f"REQ-{category.upper()[:3]}-001",
            "precondition": "애플리케이션이 정상적으로 실행된 상태",
            "테스트 절차": f"1. 애플리케이션 실행\n2. {req_text} 기능 접근\n3. 기능 동작 확인",
            "기대 결과": f"1. 기능이 정상적으로 동작함\n2. 사용자 인터페이스가 올바르게 표시됨\n3. 예상된 결과가 출력됨",
            "우선순위": priority,
            "사용자 스토리": f"사용자로서 {req_text} 기능을 사용하고 싶다",
            "원본_요구사항": req_text,
            "Figma_소스": requirement.get('source', ''),
            "노드_ID": requirement.get('node_id', ''),
            "생성일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        if test_type:
            testcase["테스트_유형"] = test_type
            if test_type.lower() in ['성능', 'performance']:
                testcase["테스트 절차"] += f"\n4. {req_text} 기능의 응답시간 측정"
                testcase["기대 결과"] += f"\n4. 응답시간이 3초 이내"
            elif test_type.lower() in ['보안', 'security']:
                testcase["테스트 절차"] += f"\n4. {req_text} 기능의 보안 취약점 확인"
                testcase["기대 결과"] += f"\n4. 보안 취약점이 발견되지 않음"
        
        return {
            "success": True,
            "testcase": testcase
        }
    
    def save_to_excel(self, test_cases: list, filename: str = None, use_template: bool = True) -> dict:
        """테스트케이스를 Excel 파일로 저장 (X Oauth.xlsx 템플릿 사용 가능)"""
        if not test_cases:
            return {
                "success": False,
                "error": "저장할 테스트케이스가 없습니다."
            }
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"mcp_figma_testcases_{timestamp}.xlsx"
        
        try:
            if use_template:
                return self._save_with_template(test_cases, filename)
            else:
                return self._save_basic_excel(test_cases, filename)
            
        except Exception as e:
            return {
                "success": False,
                "error": str(e)
            }
    
    def _save_with_template(self, test_cases: list, filename: str) -> dict:
        """X Oauth.xlsx 템플릿을 사용하여 저장"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # 템플릿 파일 경로
        template_path = "X Oauth.xlsx"
        
        try:
            # 템플릿 파일 복사
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # 기존 데이터 제거 (헤더는 유지)
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).value = None
            
            # X Oauth.xlsx 템플릿 컬럼 매핑
            template_columns = {
                'domain': 'A',
                'section': 'B', 
                'Component': 'C',
                'Feature': 'D',
                'title': 'E',
                'precondition': 'F',
                'Test step': 'G',
                'Expected-results': 'H',
                'Priority': 'I',
                'type': 'J',
                'comment': 'K',
                'android Result': 'L',
                'ios Result': 'M'
            }
            
            # 데이터 매핑 및 변환
            converted_cases = []
            for i, case in enumerate(test_cases):
                converted_case = self._convert_to_template_format(case, i)
                converted_cases.append(converted_case)
            
            # 데이터 입력
            for row_idx, case in enumerate(converted_cases, start=2):
                # domain (Social Referral v2로 고정)
                ws[f'A{row_idx}'].value = "Social Referral v2"
                ws[f'A{row_idx}'].font = Font(name='맑은 고딕')
                
                # section (플로우 타입)
                ws[f'B{row_idx}'].value = case.get('section', 'Referral Flow')
                ws[f'B{row_idx}'].font = Font(name='맑은 고딕')
                
                # Component (스크린 정보)
                ws[f'C{row_idx}'].value = case.get('component', 'UI Component')
                ws[f'C{row_idx}'].font = Font(name='맑은 고딕')
                
                # Feature (기능)
                ws[f'D{row_idx}'].value = case.get('feature', 'Core Feature')
                ws[f'D{row_idx}'].font = Font(name='맑은 고딕')
                
                # title (제목)
                ws[f'E{row_idx}'].value = case.get('title', '')
                ws[f'E{row_idx}'].font = Font(name='맑은 고딕')
                
                # precondition (사전조건)
                ws[f'F{row_idx}'].value = case.get('precondition', '')
                ws[f'F{row_idx}'].font = Font(name='맑은 고딕')
                
                # Test step (테스트 절차)
                ws[f'G{row_idx}'].value = case.get('test_step', '')
                ws[f'G{row_idx}'].font = Font(name='맑은 고딕')
                
                # Expected-results (기대 결과)
                ws[f'H{row_idx}'].value = case.get('expected_results', '')
                ws[f'H{row_idx}'].font = Font(name='맑은 고딕')
                
                # Priority (우선순위)
                ws[f'I{row_idx}'].value = case.get('priority', 'P2')
                ws[f'I{row_idx}'].font = Font(name='맑은 고딕')
                
                # type (테스트 타입)
                ws[f'J{row_idx}'].value = case.get('type', 'Functional')
                ws[f'J{row_idx}'].font = Font(name='맑은 고딕')
                
                # comment (비고)
                ws[f'K{row_idx}'].value = case.get('comment', '')
                ws[f'K{row_idx}'].font = Font(name='맑은 고딕')
                
                # android Result (결과는 비워둠)
                ws[f'L{row_idx}'].value = ''
                ws[f'L{row_idx}'].font = Font(name='맑은 고딕')
                
                # ios Result (결과는 비워둠)
                ws[f'M{row_idx}'].value = ''
                ws[f'M{row_idx}'].font = Font(name='맑은 고딕')
            
            # 파일 저장
            wb.save(filename)
            wb.close()
            
            return {
                "success": True,
                "filename": filename,
                "count": len(test_cases),
                "template_used": "X Oauth.xlsx"
            }
            
        except FileNotFoundError:
            # 템플릿 파일이 없으면 기본 형식으로 저장
            return self._save_basic_excel(test_cases, filename)
        except Exception as e:
            return {
                "success": False,
                "error": f"템플릿 적용 중 오류: {str(e)}"
            }
    
    def _convert_to_template_format(self, case: dict, index: int) -> dict:
        """테스트케이스를 X Oauth.xlsx 템플릿 형식으로 변환"""
        
        # 카테고리에서 section 추출
        category = case.get('카테고리', case.get('category', ''))
        section = category
        
        # 플로우 타입에서 component 추출
        flow_type = case.get('플로우 타입', case.get('flow_type', ''))
        if 'Internal' in flow_type:
            component = 'Internal Flow'
        elif 'External' in flow_type:
            component = 'External Flow'
        else:
            component = 'General Flow'
        
        # 스크린 정보에서 feature 추출 (업데이트: 랭킹 시스템 추가)
        screen_info = case.get('스크린', case.get('screen', ''))
        title_text = str(case.get('제목', case.get('title', '')))
        if 'KYC' in title_text:
            feature = 'KYC Process'
        elif 'Referral' in title_text:
            feature = 'Referral System'
        elif 'Login' in title_text:
            feature = 'Authentication'
        elif any(keyword in title_text.lower() for keyword in ['랭킹', 'ranking', '순위', 'rank', '리더보드', 'leaderboard', '대회', 'competition']):
            feature = 'Ranking System'
        elif any(keyword in title_text.lower() for keyword in ['vip', 'svip', '티어', 'tier', '등급', 'grade', '멤버십', 'membership', '승급', 'upgrade']):
            feature = 'VIP Tier System'
        else:
            feature = 'Core Feature'
        
        return {
            'section': section,
            'component': component,
            'feature': feature,
            'title': case.get('제목', case.get('title', f'테스트케이스 {index + 1}')),
            'precondition': case.get('사전조건', case.get('precondition', '')),
            'test_step': case.get('테스트 절차', case.get('test_step', '')),
            'expected_results': case.get('기대 결과', case.get('expected_results', '')),
            'priority': case.get('우선순위', case.get('priority', 'P2')),
            'type': case.get('테스트_유형', case.get('type', 'Functional')),
            'comment': case.get('비고', case.get('comment', ''))
        }
    
    def _save_basic_excel(self, test_cases: list, filename: str) -> dict:
        """기본 Excel 형식으로 저장 (기존 방식)"""
        df = pd.DataFrame(test_cases)
        
        column_order = [
            '카테고리', '제목', '참조 ID', 'precondition', 
            '테스트 절차', '기대 결과', '우선순위', '사용자 스토리',
            '원본_요구사항', 'Figma_소스', '노드_ID', '생성일시'
        ]
        
        existing_columns = [col for col in column_order if col in df.columns]
        df = df[existing_columns]
        
        df.to_excel(filename, index=False, sheet_name='테스트케이스')
        
        return {
            "success": True,
            "filename": filename,
            "count": len(test_cases)
        }
    
    def enhanced_figma_analysis(self, figma_url: str, include_screenshot: bool = True) -> dict:
        """향상된 Figma 분석 - 키워드 기반 + 스크린샷 유저플로우 분석"""
        
        try:
            # 1. URL 파싱
            parsed = self.parse_figma_url(figma_url)
            if not parsed.get("success"):
                return {"success": False, "error": f"URL 파싱 실패: {parsed.get('error')}"}
            
            file_id = parsed.get("file_id")
            node_id = parsed.get("node_id")
            
            # 2. Figma 데이터 가져오기
            data_result = self.fetch_figma_data(file_id, node_id)
            if not data_result.get("success"):
                return {"success": False, "error": f"데이터 가져오기 실패: {data_result.get('error')}"}
            
            figma_data = data_result.get("data", {})
            
            # 3. 기본 키워드 분석
            basic_requirements = self.extract_requirements(figma_data)
            if not basic_requirements.get("success"):
                return {"success": False, "error": "키워드 분석 실패"}
            
            # 4. 향상된 키워드 분석
            enhanced_keywords = self._analyze_enhanced_keywords(figma_data)
            
            # 5. UI 구조 분석
            ui_analysis = self._analyze_ui_structure(figma_data)
            
            # 6. 유저플로우 분석
            flow_analysis = self._analyze_user_flow(enhanced_keywords, ui_analysis)
            
            # 7. 스크린샷 분석 (옵션)
            screenshot_analysis = {}
            if include_screenshot:
                screenshot_analysis = self._analyze_screenshot(file_id, node_id)
            
            # 8. 권장사항 생성
            recommendations = self._generate_recommendations(enhanced_keywords, ui_analysis, flow_analysis)
            
            return {
                "success": True,
                "file_info": {
                    "file_id": file_id,
                    "node_id": node_id,
                    "url": figma_url
                },
                "basic_analysis": {
                    "requirements_count": basic_requirements.get("count", 0),
                    "requirements": basic_requirements.get("requirements", [])
                },
                "enhanced_analysis": {
                    "keywords": enhanced_keywords,
                    "ui_structure": ui_analysis,
                    "user_flow": flow_analysis,
                    "screenshot": screenshot_analysis if include_screenshot else None
                },
                "recommendations": recommendations,
                "summary": {
                    "total_elements": enhanced_keywords.get("total_elements", 0),
                    "ui_patterns": list(enhanced_keywords.get("detected_patterns", {}).keys()),
                    "flow_type": flow_analysis.get("primary_flow_type", "unknown"),
                    "confidence": flow_analysis.get("confidence", 0),
                    "ui_complexity": ui_analysis.get("ui_complexity", "medium")
                }
            }
            
        except Exception as e:
            return {"success": False, "error": f"향상된 분석 실패: {str(e)}"}
    
    def _analyze_enhanced_keywords(self, figma_data: dict) -> dict:
        """향상된 키워드 분석"""
        
        # UI 패턴 정의 (업데이트: 랭킹 시스템 추가)
        ui_patterns = {
            "navigation": {
                "keywords": ["nav", "menu", "tab", "breadcrumb", "back", "next", "home"],
                "flow_type": "navigation"
            },
            "authentication": {
                "keywords": ["login", "signup", "register", "signin", "oauth", "auth", "password"],
                "flow_type": "auth_flow"
            },
            "form_input": {
                "keywords": ["input", "field", "form", "textfield", "submit", "save", "cancel"],
                "flow_type": "form_interaction"
            },
            "modal_popup": {
                "keywords": ["modal", "popup", "dialog", "overlay", "confirm", "alert"],
                "flow_type": "modal_flow"
            },
            "transaction": {
                "keywords": ["buy", "sell", "trade", "order", "payment", "checkout", "confirm"],
                "flow_type": "transaction_flow"
            },
            "social": {
                "keywords": ["share", "like", "follow", "comment", "social", "connect"],
                "flow_type": "social_interaction"
            },
            "settings": {
                "keywords": ["settings", "preferences", "profile", "account", "config"],
                "flow_type": "settings_flow"
            },
            "ranking_system": {
                "keywords": ["ranking", "rank", "leaderboard", "position", "competition", "1st", "2nd", "3rd", "순위", "랭킹", "리더보드", "대회", "경쟁"],
                "flow_type": "ranking_competition"
            },
            "vip_tier_system": {
                "keywords": ["vip", "svip", "tier", "grade", "membership", "premium", "upgrade", "benefit", "privilege", "티어", "등급", "멤버십", "승급", "혜택"],
                "flow_type": "vip_membership"
            }
        }
        
        texts = []
        names = []
        
        def traverse_nodes(nodes, depth=0):
            if isinstance(nodes, list):
                for node in nodes:
                    traverse_nodes(node, depth)
            elif isinstance(nodes, dict):
                node_type = nodes.get('type')
                node_name = nodes.get('name', '')
                
                if node_type == 'TEXT' and 'characters' in nodes:
                    text = nodes['characters'].strip()
                    if text:
                        texts.append({"text": text, "depth": depth})
                
                if node_type in ['FRAME', 'COMPONENT', 'INSTANCE'] and node_name:
                    names.append({"name": node_name, "type": node_type.lower(), "depth": depth})
                
                if 'children' in nodes:
                    traverse_nodes(nodes['children'], depth + 1)
        
        traverse_nodes(figma_data.get('document', {}).get('children', []))
        
        # 모든 텍스트 결합
        all_text = " ".join([t["text"] for t in texts] + [n["name"] for n in names])
        
        # UI 패턴 매칭
        detected_patterns = {}
        for pattern_name, pattern_info in ui_patterns.items():
            keywords = pattern_info["keywords"]
            matches = sum(1 for keyword in keywords if keyword.lower() in all_text.lower())
            if matches > 0:
                detected_patterns[pattern_name] = {
                    "matches": matches,
                    "flow_type": pattern_info["flow_type"],
                    "confidence": min(matches * 20, 100)
                }
        
        return {
            "texts": texts,
            "names": names,
            "detected_patterns": detected_patterns,
            "total_elements": len(texts) + len(names)
        }
    
    def _analyze_ui_structure(self, figma_data: dict) -> dict:
        """UI 구조 분석"""
        
        ui_elements = {
            "buttons": [],
            "inputs": [],
            "navigation": [],
            "containers": []
        }
        
        layout_info = {
            "depth_levels": 0,
            "max_children": 0,
            "component_count": 0
        }
        
        def analyze_node(node, depth=0):
            if isinstance(node, dict):
                node_type = node.get('type', '')
                node_name = node.get('name', '').lower()
                
                layout_info["depth_levels"] = max(layout_info["depth_levels"], depth)
                
                # UI 요소 분류
                if 'button' in node_name or 'btn' in node_name:
                    ui_elements["buttons"].append({"name": node.get('name', ''), "depth": depth})
                elif any(keyword in node_name for keyword in ['input', 'field', 'textfield']):
                    ui_elements["inputs"].append({"name": node.get('name', ''), "depth": depth})
                elif any(keyword in node_name for keyword in ['nav', 'menu', 'tab']):
                    ui_elements["navigation"].append({"name": node.get('name', ''), "depth": depth})
                elif node_type in ['FRAME', 'GROUP']:
                    ui_elements["containers"].append({"name": node.get('name', ''), "depth": depth})
                
                if node_type in ['COMPONENT', 'INSTANCE']:
                    layout_info["component_count"] += 1
                
                children = node.get('children', [])
                if children:
                    layout_info["max_children"] = max(layout_info["max_children"], len(children))
                    for child in children:
                        analyze_node(child, depth + 1)
        
        for child in figma_data.get('document', {}).get('children', []):
            analyze_node(child)
        
        # UI 복잡도 계산
        total_elements = sum(len(elements) for elements in ui_elements.values())
        complexity_score = total_elements + layout_info["depth_levels"] * 2 + layout_info["component_count"]
        
        if complexity_score < 20:
            ui_complexity = "low"
        elif complexity_score < 50:
            ui_complexity = "medium"
        else:
            ui_complexity = "high"
        
        return {
            "ui_elements": ui_elements,
            "layout_info": layout_info,
            "ui_complexity": ui_complexity
        }
    
    def _analyze_user_flow(self, keyword_analysis: dict, ui_analysis: dict) -> dict:
        """유저플로우 분석"""
        
        detected_patterns = keyword_analysis.get("detected_patterns", {})
        ui_elements = ui_analysis.get("ui_elements", {})
        
        # 플로우 단계 추론
        flow_steps = []
        
        if "authentication" in detected_patterns:
            flow_steps.append("사용자 인증")
        else:
            flow_steps.append("화면 진입")
        
        if ui_elements.get("inputs"):
            flow_steps.append("정보 입력")
        
        if ui_elements.get("buttons"):
            button_count = len(ui_elements["buttons"])
            if button_count == 1:
                flow_steps.append("액션 실행")
            else:
                flow_steps.append("옵션 선택")
        
        flow_steps.append("결과 확인")
        
        # 주요 플로우 타입 결정
        primary_flow_type = "general"
        max_confidence = 0
        
        for pattern_name, pattern_info in detected_patterns.items():
            if pattern_info["confidence"] > max_confidence:
                max_confidence = pattern_info["confidence"]
                primary_flow_type = pattern_info["flow_type"]
        
        return {
            "flow_steps": flow_steps,
            "primary_flow_type": primary_flow_type,
            "confidence": max_confidence,
            "complexity": ui_analysis.get("ui_complexity", "medium")
        }
    
    def _analyze_screenshot(self, file_id: str, node_id: str = None) -> dict:
        """스크린샷 분석 (간단버전)"""
        
        try:
            headers = {"X-Figma-Token": self.figma_token}
            
            if node_id:
                node_id_formatted = node_id.replace('-', ':')
                image_url = f"https://api.figma.com/v1/images/{file_id}?ids={node_id_formatted}&format=png&scale=1"
            else:
                return {"success": False, "error": "노드 ID 필요"}
            
            response = requests.get(image_url, headers=headers)
            response.raise_for_status()
            
            image_data = response.json()
            
            if 'images' in image_data and image_data['images']:
                image_urls = list(image_data['images'].values())
                if image_urls and image_urls[0]:
                    image_response = requests.get(image_urls[0])
                    image_response.raise_for_status()
                    
                    image_size = len(image_response.content)
                    
                    return {
                        "success": True,
                        "image_url": image_urls[0],
                        "image_size": image_size,
                        "complexity": "high" if image_size > 500000 else "medium" if image_size > 100000 else "low"
                    }
            
            return {"success": False, "error": "이미지 생성 실패"}
            
        except Exception as e:
            return {"success": False, "error": f"스크린샷 분석 오류: {str(e)}"}
    
    def _generate_recommendations(self, keyword_analysis: dict, ui_analysis: dict, flow_analysis: dict) -> dict:
        """권장사항 생성"""
        
        recommendations = {
            "ui_improvements": [],
            "testing_priorities": [],
            "user_experience": []
        }
        
        ui_complexity = ui_analysis.get("ui_complexity", "medium")
        detected_patterns = keyword_analysis.get("detected_patterns", {})
        
        # UI 개선 권장사항
        if ui_complexity == "high":
            recommendations["ui_improvements"].append("UI 복잡도 단순화 필요")
        
        button_count = len(ui_analysis.get("ui_elements", {}).get("buttons", []))
        if button_count > 5:
            recommendations["ui_improvements"].append("주요 액션 버튼 우선순위 명확화")
        
        # 테스트 우선순위
        if "authentication" in detected_patterns:
            recommendations["testing_priorities"].append("인증 플로우 테스트 우선 실행")
        if "transaction" in detected_patterns:
            recommendations["testing_priorities"].append("거래 기능 보안 테스트 필수")
        if "form_input" in detected_patterns:
            recommendations["testing_priorities"].append("입력 유효성 검사 테스트")
        
        # 사용자 경험
        if len(flow_analysis.get("flow_steps", [])) > 5:
            recommendations["user_experience"].append("사용자 플로우 단계 단순화 고려")
        
        return recommendations

# MCP 서버 설정
if MCP_AVAILABLE:
    server = Server("figma-testcase-generator")
    
    try:
        figma_server = FigmaMCPServer()
    except ValueError as e:
        print(f"❌ 서버 초기화 실패: {e}")
        sys.exit(1)

    @server.list_tools()
    async def handle_list_tools() -> list[Tool]:
        """사용 가능한 도구 목록 반환"""
        return [
            Tool(
                name="parse_figma_url",
                description="Figma URL에서 파일 ID와 노드 ID를 추출합니다.",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "figma_url": {
                            "type": "string",
                            "description": "분석할 Figma URL"
                        }
                    },
                    "required": ["figma_url"]
                }
            ),
            Tool(
                name="fetch_figma_data",
                description="Figma API에서 파일 데이터를 가져옵니다.",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "file_id": {
                            "type": "string",
                            "description": "Figma 파일 ID"
                        },
                        "node_id": {
                            "type": "string",
                            "description": "특정 노드 ID (선택사항)"
                        }
                    },
                    "required": ["file_id"]
                }
            ),
            Tool(
                name="extract_requirements",
                description="Figma 데이터에서 요구사항을 추출합니다.",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "figma_data": {
                            "type": "object",
                            "description": "Figma API에서 가져온 데이터"
                        }
                    },
                    "required": ["figma_data"]
                }
            ),
            Tool(
                name="generate_testcase",
                description="요구사항으로부터 테스트케이스를 생성합니다.",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "requirement": {
                            "type": "object",
                            "description": "요구사항 객체"
                        },
                        "test_type": {
                            "type": "string",
                            "description": "테스트 유형 (예: UI테스트, 성능테스트)"
                        }
                    },
                    "required": ["requirement"]
                }
            ),
            Tool(
                name="save_to_excel",
                description="테스트케이스를 Excel 파일로 저장합니다. X Oauth.xlsx 템플릿 사용 가능.",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "test_cases": {
                            "type": "array",
                            "description": "저장할 테스트케이스 목록"
                        },
                        "filename": {
                            "type": "string",
                            "description": "저장할 파일명 (선택사항)"
                        },
                        "use_template": {
                            "type": "boolean",
                            "description": "X Oauth.xlsx 템플릿 사용 여부 (기본값: true)",
                            "default": True
                        }
                    },
                    "required": ["test_cases"]
                }
            ),
            Tool(
                name="process_figma_link",
                description="Figma 링크를 입력받아 전체 프로세스를 실행합니다.",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "figma_url": {
                            "type": "string",
                            "description": "처리할 Figma URL"
                        },
                        "test_type": {
                            "type": "string",
                            "description": "테스트 유형 (선택사항)"
                        }
                    },
                    "required": ["figma_url"]
                }
            ),
            Tool(
                name="enhanced_figma_analysis",
                description="향상된 Figma 분석 - 키워드 기반 + 스크린샷 유저플로우 분석을 수행합니다.",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "figma_url": {
                            "type": "string",
                            "description": "분석할 Figma URL"
                        },
                        "include_screenshot": {
                            "type": "boolean",
                            "description": "스크린샷 분석 포함 여부 (기본값: true)",
                            "default": True
                        }
                    },
                    "required": ["figma_url"]
                }
            )
        ]

    @server.call_tool()
    async def handle_call_tool(name: str, arguments: dict) -> list[TextContent]:
        """도구 호출 처리"""
        try:
            if name == "parse_figma_url":
                result = figma_server.parse_figma_url(arguments["figma_url"])
                
            elif name == "fetch_figma_data":
                result = figma_server.fetch_figma_data(
                    arguments["file_id"], 
                    arguments.get("node_id")
                )
                
            elif name == "extract_requirements":
                result = figma_server.extract_requirements(arguments["figma_data"])
                
            elif name == "generate_testcase":
                result = figma_server.generate_testcase_structure(
                    arguments["requirement"],
                    arguments.get("test_type")
                )
                
            elif name == "save_to_excel":
                result = figma_server.save_to_excel(
                    arguments["test_cases"],
                    arguments.get("filename"),
                    arguments.get("use_template", True)
                )
                
            elif name == "process_figma_link":
                # 전체 프로세스 실행
                figma_url = arguments["figma_url"]
                test_type = arguments.get("test_type")
                
                # 1. URL 파싱
                url_result = figma_server.parse_figma_url(figma_url)
                if not url_result["success"]:
                    result = {"success": False, "error": "URL 파싱 실패"}
                else:
                    # 2. Figma 데이터 가져오기
                    data_result = figma_server.fetch_figma_data(
                        url_result["file_id"], 
                        url_result.get("node_id")
                    )
                    if not data_result["success"]:
                        result = {"success": False, "error": "Figma 데이터 가져오기 실패"}
                    else:
                        # 3. 요구사항 추출
                        req_result = figma_server.extract_requirements(data_result["data"])
                        if not req_result["success"]:
                            result = {"success": False, "error": "요구사항 추출 실패"}
                        else:
                            # 4. 테스트케이스 생성
                            test_cases = []
                            for req in req_result["requirements"]:
                                tc_result = figma_server.generate_testcase_structure(req, test_type)
                                if tc_result["success"]:
                                    tc_result["testcase"]["Figma_URL"] = figma_url
                                    test_cases.append(tc_result["testcase"])
                            
                            # 5. Excel 저장 (템플릿 사용)
                            if test_cases:
                                save_result = figma_server.save_to_excel(test_cases, use_template=True)
                                result = {
                                    "success": True,
                                    "requirements_count": len(req_result["requirements"]),
                                    "testcases_count": len(test_cases),
                                    "filename": save_result.get("filename"),
                                    "template_used": save_result.get("template_used"),
                                    "figma_url": figma_url
                                }
                            else:
                                result = {"success": False, "error": "테스트케이스 생성 실패"}
                                
            elif name == "enhanced_figma_analysis":
                # 향상된 Figma 분석 실행
                figma_url = arguments["figma_url"]
                include_screenshot = arguments.get("include_screenshot", True)
                
                result = figma_server.enhanced_figma_analysis(figma_url, include_screenshot)
                
            else:
                result = {"success": False, "error": f"알 수 없는 도구: {name}"}
                
            return [TextContent(type="text", text=json.dumps(result, ensure_ascii=False, indent=2))]
            
        except Exception as e:
            error_result = {"success": False, "error": str(e)}
            return [TextContent(type="text", text=json.dumps(error_result, ensure_ascii=False, indent=2))]

    async def main():
        """MCP 서버 실행"""
        try:
            from mcp.server.stdio import stdio_server
            
            async with stdio_server() as (read_stream, write_stream):
                await server.run(
                    read_stream,
                    write_stream,
                    InitializationOptions(
                        server_name="figma-testcase-generator",
                        server_version="1.0.0",
                        capabilities=server.get_capabilities(
                            notification_options=NotificationOptions(),
                            experimental_capabilities={},
                        ),
                    ),
                )
        except Exception as e:
            print(f"❌ MCP 서버 실행 오류: {e}")

if __name__ == "__main__":
    if MCP_AVAILABLE:
        asyncio.run(main())
    else:
        print("MCP 서버를 실행하려면 'pip install mcp' 명령어로 설치해주세요.") 