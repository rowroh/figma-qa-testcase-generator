#!/usr/bin/env python3
"""
테스트케이스 생성 엔진
- 분석 결과 기반 테스트케이스 자동 생성
- 다양한 출력 형식 지원 (Excel, TestRail, JSON)
- 시나리오 기반 커스터마이징
"""

import json
import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional, Any
from ..analyzers.figma_analyzer import FigmaAnalyzer
from ..utils.rules_config import RulesConfig, load_rules_config, normalize_testcase_fields
import os

class TestCaseGenerator:
    """테스트케이스 생성기"""
    
    def __init__(self, rules: Optional[RulesConfig] = None, rules_path: Optional[str] = None):
        """초기화"""
        self.rules: RulesConfig = rules or load_rules_config(rules_path)

        # 테스트케이스 템플릿
        # NOTE: 사용자 룰세팅의 컬럼 스키마를 기본으로 사용 (test_step, web_result/app_result 등)
        self.testcase_template = {col: "" for col in self.rules.output_columns}
        self.testcase_template.update({
            "priority": self.rules.priority_default,
            "type": "Functional",
        })
        
        # 우선순위 매핑
        self.priority_mapping = {
            "critical": "P1",
            "high": "P2", 
            "medium": "P3",
            "low": "P4"
        }
        
        # 테스트 타입 정의
        self.test_types = {
            "functional": "Functional",
            "ui": "UI",
            "security": "Security",
            "performance": "Performance",
            "accessibility": "Accessibility",
            "usability": "Usability"
        }
    
    def generate_from_analysis(self, analysis_result: Dict[str, Any], 
                             custom_scenarios: Optional[List[Dict]] = None) -> List[Dict]:
        """
        Figma 분석 결과를 기반으로 테스트케이스 생성
        
        Args:
            analysis_result: FigmaAnalyzer.enhanced_analysis() 결과
            custom_scenarios: 커스텀 시나리오 (선택사항)
        
        Returns:
            List[Dict]: 생성된 테스트케이스 목록
        """
        if not analysis_result.get("success"):
            raise ValueError(f"분석 결과 오류: {analysis_result.get('error')}")
        
        testcases = []
        
        # 분석 결과에서 정보 추출
        enhanced_analysis = analysis_result.get("enhanced_analysis", {})
        keywords = enhanced_analysis.get("keywords", {})
        ui_structure = enhanced_analysis.get("ui_structure", {})
        user_flow = enhanced_analysis.get("user_flow", {})
        recommendations = analysis_result.get("recommendations", {})
        
        detected_patterns = keywords.get("detected_patterns", {})
        ui_elements = ui_structure.get("ui_elements", {})
        
        # 1. UI 패턴 기반 테스트케이스 생성
        for pattern_name, pattern_info in detected_patterns.items():
            pattern_testcases = self._generate_pattern_testcases(
                pattern_name, pattern_info, ui_elements
            )
            testcases.extend(pattern_testcases)
        
        # 2. 유저플로우 기반 테스트케이스 생성
        flow_testcases = self._generate_flow_testcases(user_flow, ui_elements)
        testcases.extend(flow_testcases)
        
        # 3. UI 요소 기반 테스트케이스 생성
        ui_testcases = self._generate_ui_testcases(ui_elements, ui_structure)
        testcases.extend(ui_testcases)
        
        # 4. 권장사항 기반 테스트케이스 생성
        recommendation_testcases = self._generate_recommendation_testcases(recommendations)
        testcases.extend(recommendation_testcases)

        # 4.5 룰 기반 기본 커버리지 보강 (접근성/사용성/엣지/네거티브/크로스플랫폼)
        rule_coverage_tests = self._generate_rule_coverage_testcases(user_flow, ui_structure, ui_elements)
        testcases.extend(rule_coverage_tests)
        
        # 5. 커스텀 시나리오 추가
        if custom_scenarios:
            custom_testcases = self._generate_custom_testcases(custom_scenarios)
            testcases.extend(custom_testcases)
        
        # 6. 우선순위 조정 및 중복 제거
        testcases = self._optimize_testcases(testcases)
        
        # 7. 필드 정규화 (test_steps -> test_step 등)
        testcases = [normalize_testcase_fields(tc, self.rules.field_aliases) for tc in testcases]
        return testcases

    def get_flow_clarification_questions(self, analysis_result: Dict[str, Any]) -> List[str]:
        """
        유저플로우가 불명확한 경우 사용자에게 확인해야 할 질문 목록 생성.
        - 현재는 heuristic 기반(신뢰도/스텝 수)
        """
        enhanced_analysis = analysis_result.get("enhanced_analysis", {})
        user_flow = enhanced_analysis.get("user_flow", {}) or {}
        confidence = float(user_flow.get("confidence", 0) or 0)
        flow_steps = user_flow.get("flow_steps", []) or []

        if confidence >= self.rules.flow_confidence_threshold and len(flow_steps) >= 3:
            return []
        return list(self.rules.flow_default_questions)
    
    def _generate_pattern_testcases(self, pattern_name: str, pattern_info: Dict, 
                                  ui_elements: Dict) -> List[Dict]:
        """UI 패턴 기반 테스트케이스 생성"""
        testcases = []
        
        if pattern_name == "authentication":
            testcases.extend(self._generate_auth_testcases(pattern_info, ui_elements))
        elif pattern_name == "form_input":
            testcases.extend(self._generate_form_testcases(pattern_info, ui_elements))
        elif pattern_name == "navigation":
            testcases.extend(self._generate_navigation_testcases(pattern_info, ui_elements))
        elif pattern_name == "modal_popup":
            testcases.extend(self._generate_modal_testcases(pattern_info, ui_elements))
        elif pattern_name == "transaction":
            testcases.extend(self._generate_transaction_testcases(pattern_info, ui_elements))
        elif pattern_name == "social":
            testcases.extend(self._generate_social_testcases(pattern_info, ui_elements))
        elif pattern_name == "settings":
            testcases.extend(self._generate_settings_testcases(pattern_info, ui_elements))
        
        return testcases
    
    def _generate_auth_testcases(self, pattern_info: Dict, ui_elements: Dict) -> List[Dict]:
        """인증 관련 테스트케이스 생성"""
        testcases = []
        
        # 기본 로그인 테스트
        testcases.append({
            **self.testcase_template,
            "domain": "authentication",
            "section": "User Authentication",
            "component": "Login",
            "feature": "Basic Login",
            "title": "정상 로그인 플로우",
            "precondition": "앱이 설치되어 있고 네트워크 연결이 활성화된 상태",
            "test_step": "1. 앱 실행\n2. 로그인 화면 확인\n3. 유효한 계정 정보 입력\n4. 로그인 버튼 클릭",
            "expected_results": "1. 앱이 정상적으로 실행됨\n2. 로그인 화면이 표시됨\n3. 계정 정보가 정상 입력됨\n4. 메인 화면으로 이동됨",
            "priority": "P1",
            "type": "Functional",
            "comment": "AI 생성 - 인증 패턴"
        })
        
        # 로그인 실패 테스트
        testcases.append({
            **self.testcase_template,
            "domain": "authentication",
            "section": "User Authentication",
            "component": "Login",
            "feature": "Login Error Handling",
            "title": "잘못된 계정 정보 로그인 시도",
            "precondition": "로그인 화면이 표시된 상태",
            "test_step": "1. 잘못된 이메일 입력\n2. 잘못된 비밀번호 입력\n3. 로그인 버튼 클릭\n4. 에러 메시지 확인",
            "expected_results": "1. 이메일이 입력됨\n2. 비밀번호가 입력됨\n3. 로그인 실패\n4. '계정 정보를 확인해주세요' 에러 메시지 표시",
            "priority": "P1",
            "type": "Functional",
            "comment": "AI 생성 - 에러 처리"
        })
        
        return testcases
    
    def _generate_form_testcases(self, pattern_info: Dict, ui_elements: Dict) -> List[Dict]:
        """폼 입력 관련 테스트케이스 생성"""
        testcases = []
        
        input_count = len(ui_elements.get("inputs", []))
        
        if input_count > 0:
            # 폼 유효성 검사 테스트
            testcases.append({
                **self.testcase_template,
                "domain": "form",
                "section": "Form Input",
                "component": "Input Validation",
                "feature": "Form Validation",
                "title": "필수 입력 필드 유효성 검사",
                "precondition": f"입력 폼이 표시된 상태 (총 {input_count}개 필드)",
                "test_step": "1. 필수 필드를 빈 상태로 두고 저장 시도\n2. 에러 메시지 확인\n3. 필수 필드 입력 후 저장\n4. 저장 완료 확인",
                "expected_results": "1. 저장 실패\n2. '필수 항목을 입력하세요' 에러 메시지\n3. 정상 저장 시도\n4. 저장 완료 메시지 표시",
                "priority": "P1",
                "type": "Functional",
                "comment": "AI 생성 - 폼 검증"
            })
        
        return testcases
    
    def _generate_navigation_testcases(self, pattern_info: Dict, ui_elements: Dict) -> List[Dict]:
        """네비게이션 관련 테스트케이스 생성"""
        testcases = []
        
        nav_count = len(ui_elements.get("navigation", []))
        
        if nav_count > 0:
            testcases.append({
                **self.testcase_template,
                "domain": "navigation",
                "section": "Navigation",
                "component": "Menu Navigation",
                "feature": "Basic Navigation",
                "title": "메뉴 네비게이션 기본 동작",
                "precondition": f"메인 화면에 네비게이션 메뉴 표시된 상태 (총 {nav_count}개 메뉴)",
                "test_step": "1. 각 메뉴 항목 클릭\n2. 해당 페이지로 이동 확인\n3. 뒤로가기 버튼 동작 확인\n4. 메뉴 선택 상태 표시 확인",
                "expected_results": "1. 메뉴 클릭이 정상 동작함\n2. 해당 페이지로 정확히 이동됨\n3. 뒤로가기가 정상 동작함\n4. 현재 선택된 메뉴가 하이라이트됨",
                "priority": "P1",
                "type": "Functional",
                "comment": "AI 생성 - 네비게이션"
            })
        
        return testcases
    
    def _generate_modal_testcases(self, pattern_info: Dict, ui_elements: Dict) -> List[Dict]:
        """모달/팝업 관련 테스트케이스 생성"""
        testcases = []
        
        testcases.append({
            **self.testcase_template,
            "domain": "ui",
            "section": "Modal",
            "component": "Popup Dialog",
            "feature": "Modal Interaction", 
            "title": "모달 팝업 기본 동작",
            "precondition": "모달을 띄울 수 있는 액션이 있는 화면",
            "test_step": "1. 모달 트리거 액션 실행\n2. 모달 팝업 표시 확인\n3. 모달 내 버튼 동작 확인\n4. 모달 닫기 동작 확인",
            "expected_results": "1. 액션이 정상 실행됨\n2. 모달이 중앙에 표시됨\n3. 모달 내 버튼이 정상 동작함\n4. 확인/취소로 모달이 닫힘",
            "priority": "P2",
            "type": "UI",
            "comment": "AI 생성 - 모달 UI"
        })
        
        return testcases
    
    def _generate_transaction_testcases(self, pattern_info: Dict, ui_elements: Dict) -> List[Dict]:
        """거래 관련 테스트케이스 생성"""
        testcases = []
        
        testcases.append({
            **self.testcase_template,
            "domain": "transaction",
            "section": "Trading",
            "component": "Order Execution",
            "feature": "Basic Trading",
            "title": "기본 거래 주문 실행",
            "precondition": "로그인된 상태이고 거래 가능한 자산이 있음",
            "test_step": "1. 거래 화면 진입\n2. 거래 종목 선택\n3. 주문 정보 입력\n4. 주문 실행\n5. 주문 완료 확인",
            "expected_results": "1. 거래 화면이 정상 로드됨\n2. 종목이 정상 선택됨\n3. 주문 정보가 정상 입력됨\n4. 주문이 정상 실행됨\n5. 주문 완료 메시지 표시",
            "priority": "P1",
            "type": "Functional",
            "comment": "AI 생성 - 거래 기능"
        })
        
        return testcases
    
    def _generate_social_testcases(self, pattern_info: Dict, ui_elements: Dict) -> List[Dict]:
        """소셜 관련 테스트케이스 생성"""
        testcases = []
        
        testcases.append({
            **self.testcase_template,
            "domain": "social",
            "section": "Social Integration",
            "component": "Social Connect",
            "feature": "Social Login",
            "title": "소셜 계정 연동",
            "precondition": "소셜 연동 기능이 활성화된 상태",
            "test_step": "1. 소셜 연동 버튼 클릭\n2. 해당 앱으로 이동 확인\n3. 인증 완료 후 앱 복귀\n4. 연동 완료 상태 확인",
            "expected_results": "1. 소셜 앱으로 정상 이동됨\n2. OAuth 인증 화면 표시됨\n3. 앱으로 정상 복귀됨\n4. 연동 완료 상태로 표시됨",
            "priority": "P1", 
            "type": "Functional",
            "comment": "AI 생성 - 소셜 연동"
        })
        
        return testcases
    
    def _generate_settings_testcases(self, pattern_info: Dict, ui_elements: Dict) -> List[Dict]:
        """설정 관련 테스트케이스 생성"""
        testcases = []
        
        testcases.append({
            **self.testcase_template,
            "domain": "settings",
            "section": "User Settings",
            "component": "Profile Settings",
            "feature": "Profile Management",
            "title": "프로필 정보 수정",
            "precondition": "프로필 설정 화면에 진입한 상태",
            "test_step": "1. 기존 프로필 정보 확인\n2. 수정 가능한 필드 편집\n3. 저장 버튼 클릭\n4. 변경사항 적용 확인",
            "expected_results": "1. 기존 정보가 정상 표시됨\n2. 필드 편집이 정상 동작함\n3. 저장이 정상 처리됨\n4. 변경사항이 즉시 반영됨",
            "priority": "P2",
            "type": "Functional",
            "comment": "AI 생성 - 설정 관리"
        })
        
        return testcases
    
    def _generate_flow_testcases(self, user_flow: Dict, ui_elements: Dict) -> List[Dict]:
        """유저플로우 기반 테스트케이스 생성"""
        testcases = []
        
        flow_steps = user_flow.get("flow_steps", [])
        primary_flow_type = user_flow.get("primary_flow_type", "general")
        
        if len(flow_steps) > 2:
            testcases.append({
                **self.testcase_template,
                "domain": "user_flow",
                "section": "User Journey",
                "component": "End-to-End Flow",
                "feature": f"{primary_flow_type.title()} Flow",
                "title": f"전체 {primary_flow_type} 플로우 검증",
                "precondition": "앱이 정상 실행된 상태",
                "test_step": "\n".join([f"{i+1}. {step}" for i, step in enumerate(flow_steps)]),
                "expected_results": "\n".join([f"{i+1}. {step}이(가) 정상 완료됨" for i, step in enumerate(flow_steps)]),
                "priority": "P1",
                "type": "Functional",
                "comment": f"AI 생성 - {primary_flow_type} 플로우"
            })
        
        return testcases
    
    def _generate_ui_testcases(self, ui_elements: Dict, ui_structure: Dict) -> List[Dict]:
        """UI 요소 기반 테스트케이스 생성"""
        testcases = []
        
        button_count = len(ui_elements.get("buttons", []))
        ui_complexity = ui_structure.get("ui_complexity", "medium")
        
        # 버튼 인터랙션 테스트
        if button_count > 0:
            testcases.append({
                **self.testcase_template,
                "domain": "ui",
                "section": "UI Elements",
                "component": "Button Interaction",
                "feature": "Button Functionality",
                "title": "버튼 인터랙션 기본 동작",
                "precondition": f"UI에 {button_count}개의 버튼이 표시된 상태",
                "test_step": "1. 각 버튼의 표시 상태 확인\n2. 버튼 클릭 동작 확인\n3. 비활성화 상태 버튼 확인\n4. 버튼 피드백 확인",
                "expected_results": "1. 모든 버튼이 정상 표시됨\n2. 클릭 시 해당 액션 실행됨\n3. 비활성화 버튼은 클릭 불가\n4. 클릭 시 시각적 피드백 제공됨",
                "priority": "P2",
                "type": "UI",
                "comment": "AI 생성 - UI 요소"
            })
        
        # UI 복잡도에 따른 테스트
        if ui_complexity == "high":
            testcases.append({
                **self.testcase_template,
                "domain": "ui",
                "section": "UI Performance",
                "component": "Complex UI",
                "feature": "UI Responsiveness",
                "title": "복잡한 UI 반응성 테스트",
                "precondition": "고복잡도 UI 화면이 로드된 상태",
                "test_step": "1. UI 로딩 시간 측정\n2. 스크롤 성능 확인\n3. 다중 인터랙션 동시 실행\n4. 메모리 사용량 확인",
                "expected_results": "1. 3초 이내 로딩 완료\n2. 스크롤이 부드럽게 동작함\n3. 인터랙션이 지연되지 않음\n4. 메모리 사용량이 적정 수준 유지",
                "priority": "P2",
                "type": "Performance",
                "comment": "AI 생성 - 성능 테스트"
            })
        
        return testcases
    
    def _generate_recommendation_testcases(self, recommendations: Dict) -> List[Dict]:
        """권장사항 기반 테스트케이스 생성"""
        testcases = []
        
        testing_priorities = recommendations.get("testing_priorities", [])
        
        for priority in testing_priorities:
            if "보안" in priority:
                testcases.append({
                    **self.testcase_template,
                    "domain": "security",
                    "section": "Security",
                    "component": "Security Test",
                    "feature": "Security Validation",
                    "title": "보안 기능 검증",
                    "precondition": "보안이 적용되어야 하는 기능",
                    "test_step": "1. 인증되지 않은 접근 시도\n2. 권한 없는 액션 실행 시도\n3. 보안 에러 처리 확인",
                    "expected_results": "1. 접근이 차단됨\n2. 권한 오류 메시지 표시\n3. 적절한 보안 처리가 수행됨",
                    "priority": "P1",
                    "type": "Security",
                    "comment": f"AI 생성 - {priority}"
                })
        
        return testcases
    
    def _generate_custom_testcases(self, custom_scenarios: List[Dict]) -> List[Dict]:
        """커스텀 시나리오 기반 테스트케이스 생성"""
        testcases = []
        
        for scenario in custom_scenarios:
            testcase = {**self.testcase_template}
            testcase.update(scenario)
            testcase["comment"] = "사용자 정의 시나리오"
            testcases.append(testcase)
        
        return testcases
    
    def _optimize_testcases(self, testcases: List[Dict]) -> List[Dict]:
        """테스트케이스 최적화 (중복 제거, 우선순위 조정)"""
        # 제목 기반 중복 제거
        seen_titles = set()
        unique_testcases = []
        
        for testcase in testcases:
            title = testcase.get("title", "")
            if title not in seen_titles:
                seen_titles.add(title)
                unique_testcases.append(testcase)
        
        # 우선순위별 정렬 (P1 > P2 > P3 > P4)
        priority_order = {"P1": 1, "P2": 2, "P3": 3, "P4": 4}
        unique_testcases.sort(key=lambda x: priority_order.get(x.get("priority", "P4"), 4))
        
        return unique_testcases
    
    def generate_scenarios(self, feature_config: Dict[str, Any]) -> List[Dict]:
        """시나리오 설정 기반 테스트케이스 생성"""
        feature_name = feature_config.get("feature_name", "Unknown Feature")
        priority = feature_config.get("priority", "P2") 
        scenarios = feature_config.get("scenarios", [])
        
        testcases = []
        
        for i, scenario in enumerate(scenarios, 1):
            testcase = {
                **self.testcase_template,
                "feature": feature_name,
                "title": f"{feature_name} - {scenario}",
                "priority": priority,
                "comment": "시나리오 기반 생성"
            }
            testcases.append(testcase)
        
        return testcases
    
    def identify_missing_tests(self, existing_tests: List[Dict], 
                             analysis_result: Dict[str, Any]) -> List[Dict]:
        """기존 테스트와 분석 결과를 비교하여 누락된 테스트 식별"""
        # 분석 결과에서 생성된 테스트케이스
        generated_tests = self.generate_from_analysis(analysis_result)
        
        # 기존 테스트 제목들
        existing_titles = {test.get("title", "") for test in existing_tests}
        
        # 누락된 테스트케이스 필터링
        missing_tests = [
            test for test in generated_tests 
            if test.get("title", "") not in existing_titles
        ]
        
        return missing_tests
    
    def generate_by_priority(self, analysis_result: Dict[str, Any], 
                           min_priority: str = "P1") -> List[Dict]:
        """우선순위 기반 테스트케이스 생성"""
        all_tests = self.generate_from_analysis(analysis_result)
        
        priority_filter = {"P1": 1, "P2": 2, "P3": 3, "P4": 4}
        min_level = priority_filter.get(min_priority, 2)
        
        filtered_tests = [
            test for test in all_tests
            if priority_filter.get(test.get("priority", "P4"), 4) <= min_level
        ]
        
        return filtered_tests
    
    def save_to_excel(self, testcases: List[Dict], filename: str):
        """Excel 형식으로 저장"""
        # 1) 룰 기반 정규화 + 컬럼 정렬
        normalized = [normalize_testcase_fields(tc, self.rules.field_aliases) for tc in testcases]
        df = pd.DataFrame(normalized)
        for col in self.rules.output_columns:
            if col not in df.columns:
                df[col] = ""
        df = df[self.rules.output_columns]

        # 2) 템플릿을 사용해 스타일 유지 (가능하면)
        template_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),  # project root
            "templates",
            "QA_Testcase_Template_WebApp.xlsx",
        )
        if os.path.exists(template_path):
            try:
                self._save_to_excel_with_template(df, filename, template_path)
                return
            except Exception:
                # 템플릿 저장 실패 시 기존 방식으로 폴백
                pass

        # 3) 폴백: DataFrame 기반 저장(스타일 일부만 적용)
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="TestCases", index=False)
            worksheet = writer.sheets["TestCases"]
            column_widths = {
                "A": 15,  # domain
                "B": 20,  # section
                "C": 20,  # component
                "D": 25,  # feature
                "E": 50,  # title
                "F": 40,  # precondition
                "G": 60,  # test_step
                "H": 60,  # expected_results
                "I": 10,  # priority
                "J": 15,  # type
                "K": 30,  # comment
                "L": 15,  # web_result
                "M": 15,  # app_result
            }
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

    def _save_to_excel_with_template(self, df: pd.DataFrame, filename: str, template_path: str) -> None:
        """
        템플릿 엑셀을 복제하여 데이터만 채움 (헤더/열너비/고정행 등 스타일 유지).
        템플릿 가정:
        - 1행: 헤더
        - 2행 이후: 데이터
        """
        import openpyxl

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # 헤더(1행) 기반 컬럼 매핑
        def _norm_header(s: str) -> str:
            return "".join(ch for ch in (s or "").strip().lower() if ch.isalnum() or ch == "_")

        header_row = 1
        header_to_col_idx: Dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            v = ws.cell(row=header_row, column=col_idx).value
            if not v:
                continue
            header_to_col_idx[_norm_header(str(v))] = col_idx

        # 템플릿에 없는 컬럼이 있으면 저장은 하되, 해당 컬럼은 스킵
        # (일관성은 rules_config.json이 관리)
        # 기존 데이터 제거 (2행부터)
        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row - 1)

        # 데이터 입력 (2행부터)
        for r_idx, row in enumerate(df.to_dict("records"), start=2):
            for key, value in row.items():
                col_idx = header_to_col_idx.get(_norm_header(key))
                if not col_idx:
                    continue
                ws.cell(row=r_idx, column=col_idx).value = value

        wb.save(filename)
        wb.close()
    
    def save_to_testrail_csv(self, testcases: List[Dict], filename: str):
        """TestRail 가져오기용 CSV 형식으로 저장"""
        # TestRail 필드로 변환
        testrail_data = []
        
        for testcase in [normalize_testcase_fields(tc, self.rules.field_aliases) for tc in testcases]:
            testrail_testcase = {
                "Section": f"{testcase.get('domain', '')}/{testcase.get('section', '')}",
                "Title": testcase.get("title", ""),
                "Type": testcase.get("type", "Functional"),
                "Priority": testcase.get("priority", "P2"),
                "Estimate": "5m",
                "References": "",
                "Preconditions": testcase.get("precondition", ""),
                "Steps": testcase.get("test_step", ""),
                "Expected Result": testcase.get("expected_results", "")
            }
            testrail_data.append(testrail_testcase)
        
        df = pd.DataFrame(testrail_data)
        df.to_csv(filename, index=False, encoding='utf-8-sig')
    
    def save_to_json(self, testcases: List[Dict], filename: str):
        """JSON 형식으로 저장"""
        normalized = [normalize_testcase_fields(tc, self.rules.field_aliases) for tc in testcases]
        output_data = {
            "metadata": {
                "generated_at": datetime.now().isoformat(),
                "total_testcases": len(normalized),
                "generator_version": "1.0.0"
            },
            "testcases": normalized
        }
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)

    def _generate_rule_coverage_testcases(self, user_flow: Dict, ui_structure: Dict, ui_elements: Dict) -> List[Dict]:
        """
        룰세팅 기반으로 항상 포함해야 하는 카테고리(접근성/사용성/네거티브/엣지/크로스플랫폼)를 보강.
        - 과도한 양산을 피하기 위해 "대표" 케이스만 추가
        """
        testcases: List[Dict] = []

        # Accessibility
        if "Accessibility" in self.rules.always_include_categories:
            testcases.append({
                **self.testcase_template,
                "domain": "accessibility",
                "section": "Accessibility",
                "component": "A11y",
                "feature": "Focus & Labels",
                "title": "접근성 레이블/포커스 이동/읽기 순서 검증",
                "precondition": "대표 화면 1개 선택(핵심 유저플로우 화면 권장)",
                "test_step": "1. 스크린리더(VoiceOver/TalkBack) 활성화\n2. 화면 요소를 순차 탐색\n3. 버튼/입력/탭의 접근성 레이블 확인\n4. 포커스 이동 순서 및 의미 단위 확인",
                "expected_results": "1. 모든 인터랙티브 요소에 의미 있는 레이블이 제공됨\n2. 포커스 이동 순서가 시각적/논리적 순서와 일치\n3. 읽기 불필요한 장식 요소는 제외됨",
                "priority": "P2",
                "type": "Accessibility",
                "comment": "룰세팅: 접근성 포함"
            })

        # Usability
        if "Usability" in self.rules.always_include_categories:
            testcases.append({
                **self.testcase_template,
                "domain": "usability",
                "section": "Usability",
                "component": "UX",
                "feature": "Microcopy & Feedback",
                "title": "사용자 피드백(로딩/성공/실패) 및 문구 가독성 검증",
                "precondition": "네트워크 요청/비동기 동작이 발생하는 대표 기능 1개",
                "test_step": "1. 요청 트리거\n2. 로딩 표시/중복 클릭 방지 확인\n3. 성공 시 토스트/상태 변화 확인\n4. 실패 시 원인 안내/재시도 동선 확인",
                "expected_results": "1. 로딩 상태가 명확히 표시되고 중복 요청이 방지됨\n2. 성공/실패 피드백이 즉시 제공됨\n3. 실패 시 사용자가 다음 액션(재시도/문의)을 선택할 수 있음",
                "priority": "P2",
                "type": "Usability",
                "comment": "룰세팅: 사용성 포함"
            })

        # Negative / Edge
        if "Negative" in self.rules.always_include_categories:
            testcases.append({
                **self.testcase_template,
                "domain": "negative",
                "section": "Error Handling",
                "component": "Network",
                "feature": "Offline/Timeout",
                "title": "네트워크 끊김/타임아웃 시 오류 처리 및 복구 동작",
                "precondition": "대표 API 호출이 있는 기능 1개",
                "test_step": "1. 요청 직후 네트워크 OFF 또는 타임아웃 유도\n2. 오류 메시지/상태 확인\n3. 재시도 버튼(또는 Pull-to-refresh) 실행\n4. 네트워크 복구 후 정상 완료 확인",
                "expected_results": "1. 오류가 명확히 안내되고 앱이 멈추지 않음\n2. 재시도 동작이 제공됨\n3. 복구 후 정상 플로우로 진행됨",
                "priority": "P1",
                "type": "Functional",
                "comment": "룰세팅: 오류 상황 커버"
            })

        if "Edge" in self.rules.always_include_categories:
            testcases.append({
                **self.testcase_template,
                "domain": "edge",
                "section": "Edge Cases",
                "component": "Boundary",
                "feature": "Input/Limit",
                "title": "경계값/최대·최소/빈 상태/초과 입력 처리",
                "precondition": "입력 또는 수량/금액 제한이 있는 대표 기능 1개",
                "test_step": "1. 최소값/최대값/초과값 입력\n2. 빈 상태에서 진행 시도\n3. 소수점/천단위 등 포맷 입력\n4. 제한 위반 시 안내 및 차단 확인",
                "expected_results": "1. 제한 위반은 차단되고 사유가 명확히 안내됨\n2. 유효 입력은 정상 처리됨\n3. 포맷/반올림 정책이 일관됨",
                "priority": "P2",
                "type": "Functional",
                "comment": "룰세팅: 엣지 케이스 커버"
            })

        # Cross-platform compatibility (web/app)
        if set(self.rules.platforms) >= {"web", "app"}:
            testcases.append({
                **self.testcase_template,
                "domain": "compatibility",
                "section": "Cross-platform",
                "component": "Web/App Parity",
                "feature": "Consistency",
                "title": "Web/App 기능/문구/상태 표시 일관성 검증",
                "precondition": "동일 기능이 Web/App에 모두 존재",
                "test_step": "1. Web에서 동일 시나리오 수행\n2. App에서 동일 시나리오 수행\n3. 입력/검증/에러 문구/상태/결과 표시 비교\n4. 차이가 있을 경우 사양/의도 여부 확인",
                "expected_results": "1. 핵심 기능 동작이 플랫폼 간 일관됨\n2. 문구/에러 처리/상태 표시가 동일하거나 사양에 의해 합리적으로 상이함\n3. 불일치 발견 시 결함 또는 기획 확인 항목으로 기록됨",
                "priority": "P2",
                "type": "Functional",
                "comment": "룰세팅: 크로스 플랫폼 고려"
            })

        return testcases

